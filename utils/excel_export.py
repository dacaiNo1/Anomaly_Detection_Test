import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

def generate_excel_report(df, highlight_flags, explanations):
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomaly Report"

    # Add full dataframe to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Apply color highlighting based on anomaly results
    for idx, flag in enumerate(highlight_flags, start=2):  # start=2 for Excel 1-based index + header
        color = {
            "red": "FF9999",
            "green": "CCFFCC",
            "blue": "CCE5FF"
        }.get(flag)
        if color:
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=idx, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Create second sheet for explanation summary
    ws_details = wb.create_sheet(title="Anomaly Details")
    explanation_counts = pd.Series(explanations).value_counts().reset_index()
    explanation_counts.columns = ["Anomaly Explanation", "Count"]
    for r in dataframe_to_rows(explanation_counts, index=False, header=True):
        ws_details.append(r)

    # Add chart to explanation sheet
    chart = BarChart()
    chart.type = "col"
    chart.title = "Anomaly Explanation Frequency"
    chart.x_axis.title = "Explanation"
    chart.y_axis.title = "Count"

    data = Reference(ws_details, min_col=2, min_row=1, max_row=ws_details.max_row)
    cats = Reference(ws_details, min_col=1, min_row=2, max_row=ws_details.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_details.add_chart(chart, "E2")

    # Save workbook to memory and return
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
